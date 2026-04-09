# -*- coding: utf-8 -*-
"""
档案管理系统路由 - 延迟加载版本
避免循环导入问题
"""
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, current_app
from flask_login import login_required, current_user
import os
from datetime import datetime
from utils import save_upload_file, allowed_file

# 延迟导入模型和处理器
def get_models():
    from archive_models import (
        ArchiveFonds, ArchiveCatalog, ArchiveVolume, 
        ArchiveFile, ArchiveBorrow, ArchiveNotification,
        ArchiveDigitizationTask
    )
    return ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, ArchiveBorrow, ArchiveNotification, ArchiveDigitizationTask

def get_digitizer():
    from archive_digitizer import archive_digitizer
    return archive_digitizer

archive_bp = Blueprint('archive', __name__, url_prefix='/archive')


# ==================== 首页 ====================

@archive_bp.route('/')
@login_required
def index():
    """档案管理首页"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    # 统计数据
    stats = {
        'fonds_count': ArchiveFonds.query.filter_by(is_active=True).count(),
        'catalog_count': ArchiveCatalog.query.count(),
        'volume_count': ArchiveVolume.query.count(),
        'file_count': ArchiveFile.query.count(),
        'digitized_count': ArchiveFile.query.filter_by(is_digitized=True).count()
    }
    
    # 最近档案
    recent_files = ArchiveFile.query.order_by(ArchiveFile.created_at.desc()).limit(10).all()
    
    # 全宗列表
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).limit(5).all()
    
    return render_template('archive/index.html', stats=stats, recent_files=recent_files, fonds_list=fonds_list)


# ==================== 档案详情 ====================


@archive_bp.route('/file/<int:file_id>')
@login_required
def file_detail(file_id):
    """档案详情页"""
    _, _, _, ArchiveFile, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    archive_file = ArchiveFile.query.get_or_404(file_id)
    
    # 获取借阅记录
    borrows = ArchiveBorrow.query.filter_by(archive_file_id=file_id).order_by(
        ArchiveBorrow.borrow_date.desc()
    ).limit(10).all()
    
    return render_template('archive/file_detail.html', file=archive_file, borrows=borrows)


# ==================== 全宗管理 ====================


@archive_bp.route('/batch_upload')
@login_required
def batch_upload():
    """批量上传页面"""
    ArchiveFonds, ArchiveCatalog, _, _, _, _, _ = get_models()
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()
    catalogs_list = ArchiveCatalog.query.order_by(ArchiveCatalog.fonds_id, ArchiveCatalog.catalog_code).all()
    return render_template(
        'archive/batch_upload.html',
        fonds_list=fonds_list,
        catalogs_list=catalogs_list
    )

@archive_bp.route('/fonds')
@login_required
def fonds_list():
    """全宗列表"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db

    fonds_list = ArchiveFonds.query.filter_by(is_active=True).order_by(ArchiveFonds.fonds_code).all()
    total_files = ArchiveFile.query.filter_by(status='active').count()

    catalog_counts = dict(
        db.session.query(ArchiveCatalog.fonds_id, db.func.count(ArchiveCatalog.id))
        .group_by(ArchiveCatalog.fonds_id)
        .all()
    )
    volume_counts = dict(
        db.session.query(ArchiveVolume.fonds_id, db.func.count(ArchiveVolume.id))
        .group_by(ArchiveVolume.fonds_id)
        .all()
    )
    file_counts = dict(
        db.session.query(ArchiveFile.fonds_id, db.func.count(ArchiveFile.id))
        .filter(ArchiveFile.status == 'active')
        .group_by(ArchiveFile.fonds_id)
        .all()
    )

    for fonds in fonds_list:
        fonds.catalog_count = catalog_counts.get(fonds.id, 0)
        fonds.volume_count = volume_counts.get(fonds.id, 0)
        fonds.file_count = file_counts.get(fonds.id, 0)

    return render_template(
        'archive/fonds_list.html',
        fonds_list=fonds_list,
        fonds=fonds_list,
        total_files=total_files
    )


@archive_bp.route('/fonds/create', methods=['POST'])
@login_required
def fonds_create():
    """创建全宗"""
    ArchiveFonds, _, _, _, _, _, _ = get_models()
    from models import db
    
    data = request.form
    try:
        fonds = ArchiveFonds(
            fonds_code=data.get('fonds_code'),
            fonds_name=data.get('fonds_name'),
            fonds_type=data.get('fonds_type', '机关'),
            description=data.get('description'),
            created_by=current_user.id
        )
        db.session.add(fonds)
        db.session.commit()
        flash('全宗创建成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'全宗创建失败：{str(e)}', 'error')
    return redirect(url_for('archive.fonds_list'))


@archive_bp.route('/fonds/<int:fonds_id>/edit', methods=['GET', 'POST'])
@login_required
def fonds_edit(fonds_id):
    """编辑全宗"""
    ArchiveFonds, _, _, _, _, _, _ = get_models()
    from models import db
    fonds = ArchiveFonds.query.get_or_404(fonds_id)
    if request.method == 'POST':
        fonds.fonds_code = request.form.get('fonds_code', fonds.fonds_code)
        fonds.fonds_name = request.form.get('fonds_name', fonds.fonds_name)
        fonds.fonds_type = request.form.get('fonds_type', fonds.fonds_type)
        fonds.description = request.form.get('description', fonds.description)
        try:
            db.session.commit()
            flash('全宗更新成功', 'success')
            return redirect(url_for('archive.fonds_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    return render_template('archive/fonds_edit.html', fonds=fonds)


@archive_bp.route('/fonds/<int:fonds_id>/delete', methods=['POST'])
@login_required
def fonds_delete(fonds_id):
    """删除全宗"""
    ArchiveFonds, ArchiveCatalog, _, ArchiveFile, _, _, _ = get_models()
    from models import db
    fonds = ArchiveFonds.query.get_or_404(fonds_id)
    has_catalogs = ArchiveCatalog.query.filter_by(fonds_id=fonds_id).first() is not None
    has_files = ArchiveFile.query.filter_by(fonds_id=fonds_id).first() is not None
    if has_catalogs or has_files:
        flash('该全宗下存在目录或档案，无法删除', 'error')
        return redirect(url_for('archive.fonds_list'))
    try:
        db.session.delete(fonds)
        db.session.commit()
        flash('全宗删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    return redirect(url_for('archive.fonds_list'))


# ==================== 目录管理 ====================

@archive_bp.route('/catalog/<int:fonds_id>')
@login_required
def catalog_list(fonds_id):
    """目录列表"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db

    fonds = ArchiveFonds.query.get_or_404(fonds_id)
    catalogs = (
        ArchiveCatalog.query
        .filter_by(fonds_id=fonds_id, parent_id=None)
        .order_by(ArchiveCatalog.catalog_code)
        .all()
    )

    volume_counts = dict(
        db.session.query(ArchiveVolume.catalog_id, db.func.count(ArchiveVolume.id))
        .group_by(ArchiveVolume.catalog_id)
        .all()
    )
    file_counts = dict(
        db.session.query(ArchiveFile.catalog_id, db.func.count(ArchiveFile.id))
        .filter(ArchiveFile.status == 'active')
        .group_by(ArchiveFile.catalog_id)
        .all()
    )

    for catalog in catalogs:
        catalog.child_count = catalog.children.count()
        catalog.volume_count = volume_counts.get(catalog.id, 0)
        catalog.file_count = file_counts.get(catalog.id, 0)

    return render_template(
        'archive/catalog_list.html',
        fonds=fonds,
        catalogs=catalogs,
        total_volumes=sum(c.volume_count for c in catalogs),
        total_files=sum(c.file_count for c in catalogs)
    )


@archive_bp.route('/catalog/create', methods=['POST'])
@login_required
def catalog_create():
    """创建目录"""
    _, ArchiveCatalog, _, _, _, _, _ = get_models()
    from models import db
    
    data = request.form
    catalog = ArchiveCatalog(
        fonds_id=data.get('fonds_id'),
        catalog_code=data.get('catalog_code'),
        catalog_name=data.get('catalog_name'),
        catalog_type=data.get('catalog_type'),
        parent_id=data.get('parent_id') or None,
        retention_period=data.get('retention_period'),
        description=data.get('description')
    )
    db.session.add(catalog)
    db.session.commit()
    flash('目录创建成功', 'success')
    return redirect(url_for('archive.catalog_list', fonds_id=data.get('fonds_id')))


@archive_bp.route('/catalog/<int:catalog_id>/edit', methods=['GET', 'POST'])
@login_required
def catalog_edit(catalog_id):
    """编辑目录"""
    _, ArchiveCatalog, _, _, _, _, _ = get_models()
    from models import db
    catalog = ArchiveCatalog.query.get_or_404(catalog_id)
    if request.method == 'POST':
        catalog.catalog_code = request.form.get('catalog_code', catalog.catalog_code)
        catalog.catalog_name = request.form.get('catalog_name', catalog.catalog_name)
        catalog.catalog_type = request.form.get('catalog_type', catalog.catalog_type)
        catalog.retention_period = request.form.get('retention_period', catalog.retention_period)
        catalog.description = request.form.get('description', catalog.description)
        try:
            db.session.commit()
            flash('目录更新成功', 'success')
            return redirect(url_for('archive.catalog_list', fonds_id=catalog.fonds_id))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    return render_template('archive/catalog_edit.html', catalog=catalog)


@archive_bp.route('/catalog/<int:catalog_id>/delete', methods=['POST'])
@login_required
def catalog_delete(catalog_id):
    """删除目录"""
    _, ArchiveCatalog, _, ArchiveFile, _, _, _ = get_models()
    from models import db
    catalog = ArchiveCatalog.query.get_or_404(catalog_id)
    fonds_id = catalog.fonds_id
    has_files = ArchiveFile.query.filter_by(catalog_id=catalog_id).first() is not None
    if has_files:
        flash('该目录下存在档案，无法删除', 'error')
        return redirect(url_for('archive.catalog_list', fonds_id=fonds_id))
    try:
        db.session.delete(catalog)
        db.session.commit()
        flash('目录删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    return redirect(url_for('archive.catalog_list', fonds_id=fonds_id))


# ==================== 案卷管理 ====================

@archive_bp.route('/volume/<int:catalog_id>')
@login_required
def volume_list(catalog_id):
    """案卷列表"""
    _, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db

    catalog = ArchiveCatalog.query.get_or_404(catalog_id)
    volumes = (
        ArchiveVolume.query
        .filter_by(catalog_id=catalog_id)
        .order_by(ArchiveVolume.volume_code)
        .all()
    )

    file_counts = dict(
        db.session.query(ArchiveFile.volume_id, db.func.count(ArchiveFile.id))
        .filter(ArchiveFile.volume_id.isnot(None), ArchiveFile.status == 'active')
        .group_by(ArchiveFile.volume_id)
        .all()
    )

    for volume in volumes:
        volume.file_count = file_counts.get(volume.id, 0)

    return render_template(
        'archive/volume_list.html',
        fonds=catalog.fonds,
        catalog=catalog,
        volumes=volumes,
        total_files=sum(v.file_count for v in volumes)
    )



@archive_bp.route('/volume/create', methods=['POST'])
@login_required
def volume_create():
    """创建案卷"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, _, _, _, _ = get_models()
    from models import db
    
    data = request.form
    volume = ArchiveVolume(
        fonds_id=data.get('fonds_id'),
        catalog_id=data.get('catalog_id'),
        volume_code=data.get('volume_code'),
        volume_title=data.get('volume_title'),
        volume_year=data.get('volume_year'),
        retention_period=data.get('retention_period'),
        security_level=data.get('security_level', '公开'),
        responsibility=data.get('responsibility'),
        storage_location=data.get('storage_location'),
        description=data.get('description'),
        created_by=current_user.id
    )
    db.session.add(volume)
    db.session.commit()
    flash('案卷创建成功', 'success')
    return redirect(url_for('archive.volume_list', catalog_id=data.get('catalog_id')))


@archive_bp.route('/volume/<int:volume_id>/edit', methods=['GET', 'POST'])
@login_required
def volume_edit(volume_id):
    """编辑案卷"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, _, _, _, _ = get_models()
    from models import db
    volume = ArchiveVolume.query.get_or_404(volume_id)
    if request.method == 'POST':
        volume.volume_code = request.form.get('volume_code', volume.volume_code)
        volume.volume_title = request.form.get('volume_title', volume.volume_title)
        volume.volume_year = request.form.get('volume_year', volume.volume_year)
        volume.retention_period = request.form.get('retention_period', volume.retention_period)
        volume.security_level = request.form.get('security_level', volume.security_level)
        volume.responsibility = request.form.get('responsibility', volume.responsibility)
        volume.storage_location = request.form.get('storage_location', volume.storage_location)
        volume.description = request.form.get('description', volume.description)
        try:
            db.session.commit()
            flash('案卷更新成功', 'success')
            return redirect(url_for('archive.volume_list', catalog_id=volume.catalog_id))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    # 查询全宗和目录对象，避免模板中直接访问 volume.fonds/volume.catalog 报错
    fonds = ArchiveFonds.query.get(volume.fonds_id)
    catalog = ArchiveCatalog.query.get(volume.catalog_id)
    return render_template('archive/volume_edit.html', volume=volume, fonds=fonds, catalog=catalog)


@archive_bp.route('/volume/<int:volume_id>/delete', methods=['POST'])
@login_required
def volume_delete(volume_id):
    """删除案卷"""
    _, _, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db
    volume = ArchiveVolume.query.get_or_404(volume_id)
    catalog_id = volume.catalog_id
    has_files = ArchiveFile.query.filter_by(volume_id=volume_id).first() is not None
    if has_files:
        flash('该案卷下存在档案，无法删除', 'error')
        return redirect(url_for('archive.volume_list', catalog_id=catalog_id))
    try:
        db.session.delete(volume)
        db.session.commit()
        flash('案卷删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    return redirect(url_for('archive.volume_list', catalog_id=catalog_id))


# ==================== 档案文件管理 ====================

@archive_bp.route('/files')
@login_required
def file_list():
    """档案文件列表"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    # 筛选条件
    fonds_id = request.args.get('fonds_id', type=int)
    catalog_id = request.args.get('catalog_id', type=int)
    year = request.args.get('year', type=int)
    archive_type = request.args.get('type')
    keyword = request.args.get('keyword')
    
    query = ArchiveFile.query.filter_by(status='active')
    
    if fonds_id:
        query = query.filter_by(fonds_id=fonds_id)
    if catalog_id:
        query = query.filter_by(catalog_id=catalog_id)
    if year:
        query = query.filter_by(file_year=year)
    if archive_type:
        query = query.filter_by(archive_type=archive_type)
    if keyword:
        search = f"%{keyword}%"
        query = query.filter(
            db.or_(
                ArchiveFile.title.like(search),
                ArchiveFile.keywords.like(search),
                ArchiveFile.responsibility.like(search)
            )
        )
    
    # 分页
    page = request.args.get('page', 1, type=int)
    per_page = 20
    pagination = query.order_by(ArchiveFile.file_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # 全宗列表（用于筛选）
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()
    
    # 统计数据
    stats = {
        'digitized': ArchiveFile.query.filter_by(status='active', is_digitized=True).count(),
        'permanent': ArchiveFile.query.filter_by(status='active', retention_period='永久').count(),
        'borrowed': ArchiveFile.query.join(ArchiveBorrow).filter(ArchiveBorrow.status == 'borrowed').count() if ArchiveBorrow else 0,
    }
    
    return render_template('archive/file_list.html', 
                         files=pagination.items,
                         pagination=pagination,
                         fonds_list=fonds_list,
                         stats=stats,
                         filters={
                             'fonds_id': fonds_id,
                             'catalog_id': catalog_id,
                             'year': year,
                             'type': archive_type,
                             'keyword': keyword
                         })


@archive_bp.route('/file/upload', methods=['GET', 'POST'])
@login_required
def file_upload():
    """档案上传（单文件）- 集成AI智能提取"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db
    
    if request.method == 'POST':
        fonds_id = request.form.get('fonds_id')
        catalog_id = request.form.get('catalog_id')
        volume_id = request.form.get('volume_id') or None
        
        if 'file' not in request.files:
            flash('未选择文件', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('未选择文件', 'error')
            return redirect(request.url)
        
        # 保存临时文件
        temp_path = save_upload_file(file, 'temp')
        if not temp_path:
            flash('文件保存失败', 'error')
            return redirect(request.url)
        
        # ===== AI 智能提取内容（复用知识库算法）=====
        try:
            from smart_knowledge import smart_kb
            ai_content = smart_kb.extract_content(temp_path)
            ai_keywords = smart_kb.generate_keywords(ai_content, top_k=10) if ai_content else []
            ai_summary = smart_kb.generate_summary(ai_content, max_length=300) if ai_content else ''
            ai_tags = smart_kb.auto_tag(ai_content) if ai_content else []
            ai_title = smart_kb.suggest_title(ai_content, file.filename) if ai_content else ''
        except Exception as e:
            ai_content = ''
            ai_keywords = []
            ai_summary = ''
            ai_tags = []
            ai_title = ''
        
        # 元数据（用户填写优先，AI填充作为兜底）
        user_title = request.form.get('title', '').strip()
        user_keywords = request.form.get('keywords', '').strip()
        user_description = request.form.get('description', '').strip()
        
        metadata = {
            'title': user_title or ai_title or os.path.splitext(file.filename)[0],
            'responsibility': request.form.get('responsibility'),
            'file_date': request.form.get('file_date'),
            'retention_period': request.form.get('retention_period'),
            'security_level': request.form.get('security_level'),
            'archive_type': request.form.get('archive_type'),
            'description': user_description or ai_summary,
            'reference_number': request.form.get('reference_number'),
            # AI提取字段
            'content_text': ai_content,
            'keywords': user_keywords or ','.join(ai_keywords),
            'tags': ','.join(ai_tags),
            'summary': ai_summary,
        }
        
        # 数字化处理
        archive_digitizer = get_digitizer()
        result = archive_digitizer.process_digitization(
            temp_path, fonds_id, catalog_id, volume_id,
            current_user.id, metadata
        )
        
        # 清理临时文件
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        if isinstance(result, ArchiveFile):
            # 补充AI提取的内容（数字化处理可能没有写入这些字段）
            if ai_content and not result.content_text:
                result.content_text = ai_content
            if ai_keywords and not result.keywords:
                result.keywords = ','.join(ai_keywords)
            if ai_summary and not result.summary:
                result.summary = ai_summary
            if ai_tags and not result.tags:
                result.tags = ','.join(ai_tags)
            db.session.commit()
            
            flash(f'档案上传成功！档号: {result.get_archive_code()}，AI已自动提取关键词和摘要', 'success')
            return redirect(url_for('archive.file_detail', file_id=result.id))
        else:
            flash(f'上传失败: {result.get("error", "未知错误")}', 'error')
            return redirect(request.url)
    
    # GET - 显示上传表单
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()
    return render_template('archive/file_upload.html', fonds_list=fonds_list)


@archive_bp.route('/file/<int:file_id>/edit', methods=['GET', 'POST'])
@login_required
def file_edit(file_id):
    """编辑档案文件"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db

    archive_file = ArchiveFile.query.get_or_404(file_id)

    if request.method == 'POST':
        from datetime import datetime
        archive_file.fonds_id = request.form.get('fonds_id', type=int)
        archive_file.catalog_id = request.form.get('catalog_id', type=int)
        archive_file.volume_id = request.form.get('volume_id', type=int) or None
        archive_file.file_code = request.form.get('file_code', archive_file.file_code)
        archive_file.title = request.form.get('title', archive_file.title)
        archive_file.responsibility = request.form.get('responsibility', archive_file.responsibility)
        file_date_str = request.form.get('file_date')
        archive_file.file_date = datetime.strptime(file_date_str, '%Y-%m-%d').date() if file_date_str else None
        archive_file.file_year = request.form.get('file_year', type=int) or None
        archive_file.archive_type = request.form.get('archive_type', archive_file.archive_type)
        archive_file.retention_period = request.form.get('retention_period', archive_file.retention_period)
        archive_file.security_level = request.form.get('security_level', archive_file.security_level)
        archive_file.reference_number = request.form.get('reference_number', archive_file.reference_number)
        archive_file.page_count = request.form.get('page_count', type=int) or None
        archive_file.storage_location = request.form.get('storage_location', archive_file.storage_location)
        archive_file.keywords = request.form.get('keywords', archive_file.keywords)
        archive_file.summary = request.form.get('summary', archive_file.summary)
        archive_file.tags = request.form.get('tags', archive_file.tags)
        archive_file.description = request.form.get('description', archive_file.description)
        try:
            db.session.commit()
            flash('档案更新成功', 'success')
            return redirect(url_for('archive.file_detail', file_id=archive_file.id))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')

    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()
    # 加载当前全宗下的所有目录，以及当前目录下的所有案卷，供前端选择器使用
    catalog_list = ArchiveCatalog.query.filter_by(fonds_id=archive_file.fonds_id).all() if archive_file.fonds_id else []
    volume_list = ArchiveVolume.query.filter_by(catalog_id=archive_file.catalog_id).all() if archive_file.catalog_id else []
    return render_template('archive/file_edit.html', file=archive_file, fonds_list=fonds_list,
                           catalog_list=catalog_list, volume_list=volume_list)


@archive_bp.route('/file/<int:file_id>/delete', methods=['POST'])
@login_required
def file_delete(file_id):
    """删除档案文件"""
    _, _, _, ArchiveFile, _, _, _ = get_models()
    from models import db

    archive_file = ArchiveFile.query.get_or_404(file_id)
    try:
        db.session.delete(archive_file)
        db.session.commit()
        flash('档案删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    return redirect(url_for('archive.file_list'))


def _do_re_extract(file_id):
    """实际提取逻辑，供内部和公开路由调用"""
    _, _, _, ArchiveFile, _, _, _ = get_models()
    archive_file = ArchiveFile.query.get_or_404(file_id)

    file_path = archive_file.file_path
    if not file_path or not os.path.exists(file_path):
        return jsonify({'error': '原始文件不存在，无法提取'}), 400

    try:
        from smart_knowledge import smart_kb
        content = smart_kb.extract_content(file_path)

        # 如果提取结果是错误标记（以 [ 开头），给出友好说明
        has_content = bool(content and not content.startswith('['))

        keywords = smart_kb.generate_keywords(content, top_k=10) if has_content else []
        summary = smart_kb.generate_summary(content, max_length=300) if has_content else ''
        tags = smart_kb.auto_tag(content) if has_content else []
        suggested_title = smart_kb.suggest_title(content, archive_file.original_filename or file_path) if has_content else ''

        return jsonify({
            'success': True,
            'has_content': has_content,
            'extract_note': content if not has_content else '',
            'title': suggested_title,
            'keywords': keywords,
            'keywords_str': ','.join(str(k) for k in keywords),
            'tags': tags,
            'tags_str': ','.join(str(t) for t in tags),
            'summary': summary,
            'word_count': len(content) if has_content else 0,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@archive_bp.route('/api/file/<int:file_id>/re-extract', methods=['POST'])
@login_required
def api_re_extract_file(file_id):
    """
    对已入库档案重新执行AI提取（编辑页"重新提取"按钮）
    读取 archive_file.file_path 中的原始文件，调用 smart_kb 提取内容/关键词/摘要/标签
    """
    return _do_re_extract(file_id)


@archive_bp.route('/api/analyze_file', methods=['POST'])
@login_required
def api_analyze_file():
    """
    档案文件预分析 API（上传前AI智能解析）
    复用知识库的 smart_kb 算法
    """
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '未选择文件'}), 400
    
    try:
        # 临时保存
        temp_path = save_upload_file(file, 'temp')
        if not temp_path:
            return jsonify({'error': '文件保存失败'}), 500
        
        # 使用知识库算法分析
        from smart_knowledge import smart_kb
        content = smart_kb.extract_content(temp_path)
        keywords = smart_kb.generate_keywords(content, top_k=10)
        summary = smart_kb.generate_summary(content, max_length=300)
        tags = smart_kb.auto_tag(content)
        suggested_title = smart_kb.suggest_title(content, file.filename)
        
        # 清理临时文件
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'title': suggested_title,
            'keywords': keywords,
            'keywords_str': ','.join(keywords),
            'tags': tags,
            'tags_str': ','.join(tags),
            'summary': summary,
            'word_count': len(content) if content else 0,
            'has_content': bool(content and not content.startswith('['))
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@archive_bp.route('/api/batch_upload', methods=['POST'])
@login_required
def api_batch_upload():
    """
    批量上传档案 API
    复用知识库的批量处理功能
    """
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db
    
    fonds_id = request.form.get('fonds_id')
    catalog_id = request.form.get('catalog_id')
    volume_id = request.form.get('volume_id') or None
    
    if not fonds_id or not catalog_id:
        return jsonify({'error': '请选择全宗和目录'}), 400
    
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未选择文件'}), 400
    
    # 保存临时文件
    temp_paths = []
    for file in files:
        if file and allowed_file(file.filename):
            temp_path = save_upload_file(file, 'temp')
            if temp_path:
                temp_paths.append(temp_path)
    
    # 批量处理
    def progress_callback(current, total, filename):
        pass
    
    archive_digitizer = get_digitizer()
    results = archive_digitizer.process_batch(
        temp_paths, fonds_id, catalog_id, volume_id,
        current_user.id, progress_callback
    )
    
    # 清理临时文件
    for path in temp_paths:
        if os.path.exists(path):
            os.remove(path)
    
    return jsonify({
        'success': len(results['success']),
        'failed': len(results['failed']),
        'errors': results['failed'],
        'archives': [
            {
                'id': a.id,
                'code': a.get_archive_code(),
                'title': a.title,
                'keywords': a.keywords,
                'tags': a.tags
            }
            for a in results['success']
        ]
    })


# ==================== 档案检索 ====================

@archive_bp.route('/search')
@login_required
def search():
    """档案检索"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    
    query = (request.args.get('q') or request.args.get('keyword') or '').strip()
    fonds_id = request.args.get('fonds_id', type=int)
    catalog_id = request.args.get('catalog_id', type=int)
    archive_type = request.args.get('type')
    year = request.args.get('year', type=int)
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()

    if not query:
        return render_template('archive/search.html', results=[], query='', fonds_list=fonds_list)

    # 智能检索
    archive_digitizer = get_digitizer()
    results = archive_digitizer.search_archives(
        query, fonds_id, catalog_id, archive_type, year
    )

    return render_template('archive/search.html',
                         results=results,
                         query=query,
                         fonds_list=fonds_list)



# ==================== 档案借阅 ====================

@archive_bp.route('/borrow/<int:file_id>', methods=['POST'])
@login_required
def borrow_request(file_id):
    """申请借阅"""
    _, _, _, ArchiveFile, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    archive_file = ArchiveFile.query.get_or_404(file_id)
    
    # 检查是否已借出
    active_borrow = ArchiveBorrow.query.filter_by(
        archive_file_id=file_id,
        status='borrowed'
    ).first()
    
    if active_borrow:
        flash('该档案已被借出', 'error')
        return redirect(url_for('archive.file_detail', file_id=file_id))
    
    try:
        borrow = ArchiveBorrow(
            archive_file_id=file_id,
            borrower_id=current_user.id,
            return_date=request.form.get('return_date'),
            purpose=request.form.get('purpose'),
            status='pending'
        )
        db.session.add(borrow)
        db.session.commit()
        flash('借阅申请已提交，等待审批', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'借阅申请提交失败：{str(e)}', 'error')
    return redirect(url_for('archive.file_detail', file_id=file_id))


@archive_bp.route('/borrow')
@login_required
def borrow():
    """借阅管理 - 重定向到我的借阅列表"""
    return redirect(url_for('archive.my_borrows'))


@archive_bp.route('/my_borrows')
@login_required
def my_borrows():
    """我的借阅"""
    _, _, _, _, ArchiveBorrow, _, _ = get_models()

    status = request.args.get('status', 'all')
    page = request.args.get('page', 1, type=int)
    per_page = 10

    base_query = ArchiveBorrow.query.filter_by(borrower_id=current_user.id).order_by(
        ArchiveBorrow.borrow_date.desc()
    )

    if status == 'active':
        filtered_query = base_query.filter(ArchiveBorrow.status == 'borrowed')
    elif status == 'returned':
        filtered_query = base_query.filter(ArchiveBorrow.status == 'returned')
    elif status == 'pending':
        filtered_query = base_query.filter(ArchiveBorrow.status == 'pending')
    elif status == 'rejected':
        filtered_query = base_query.filter(ArchiveBorrow.status == 'rejected')
    elif status == 'overdue':
        filtered_query = base_query.filter(
            ArchiveBorrow.status == 'borrowed',
            ArchiveBorrow.return_date.isnot(None),
            ArchiveBorrow.return_date < datetime.now()
        )
    else:
        filtered_query = base_query

    pagination = filtered_query.paginate(page=page, per_page=per_page, error_out=False)
    all_borrows = base_query.all()

    stats = {
        'total': len(all_borrows),
        'active': len([b for b in all_borrows if b.status == 'borrowed']),
        'pending': len([b for b in all_borrows if b.status == 'pending']),
        'returned': len([b for b in all_borrows if b.status == 'returned']),
        'rejected': len([b for b in all_borrows if b.status == 'rejected']),
        'overdue': len([b for b in all_borrows if b.is_overdue]),
    }

    return render_template(
        'archive/my_borrows.html',
        borrows=pagination.items,
        pagination=pagination,
        stats=stats,
        status=status
    )


@archive_bp.route('/borrow_manage')
@login_required
def borrow_manage():
    """借阅审批管理（管理员/经理）"""
    _, _, _, ArchiveFile, ArchiveBorrow, _, _ = get_models()
    from models import db
    from sqlalchemy import or_
    
    # 获取所有待审批和借出中的借阅申请
    status = request.args.get('status', 'pending')
    
    if status == 'pending':
        borrows = ArchiveBorrow.query.filter_by(status='pending').order_by(
            ArchiveBorrow.borrow_date.desc()
        ).all()
    elif status == 'borrowed':
        borrows = ArchiveBorrow.query.filter_by(status='borrowed').order_by(
            ArchiveBorrow.borrow_date.desc()
        ).all()
    elif status == 'returned':
        borrows = ArchiveBorrow.query.filter_by(status='returned').order_by(
            ArchiveBorrow.return_date.desc()
        ).all()
    else:
        borrows = ArchiveBorrow.query.order_by(ArchiveBorrow.borrow_date.desc()).all()
    
    # 统计数据
    stats = {
        'pending': ArchiveBorrow.query.filter_by(status='pending').count(),
        'borrowed': ArchiveBorrow.query.filter_by(status='borrowed').count(),
        'returned': ArchiveBorrow.query.filter_by(status='returned').count(),
        'total': ArchiveBorrow.query.count()
    }
    
    return render_template('archive/borrow_manage.html', borrows=borrows, stats=stats, current_status=status)


@archive_bp.route('/borrow/approve/<int:borrow_id>', methods=['POST'])
@login_required
def borrow_approve(borrow_id):
    """审批借阅申请（集成消息通知）"""
    _, _, _, _, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    borrow = ArchiveBorrow.query.get_or_404(borrow_id)
    
    if borrow.status != 'pending':
        return jsonify({'success': False, 'message': '只能审批待处理的申请'})
    
    action = request.form.get('action')
    archive_title = borrow.archive_file.title if borrow.archive_file else '档案'
    
    try:
        if action == 'approve':
            borrow.status = 'borrowed'
            borrow.approver_id = current_user.id
            borrow.approved_at = datetime.now()
            message = '借阅申请已批准'
            notif_title = '借阅申请已批准'
            notif_content = f'您对档案《{archive_title}》的借阅申请已获批准，请按时归还。'
            notif_type = 'success'
        else:
            reason = request.form.get('reason', '')
            borrow.status = 'rejected'
            borrow.approver_id = current_user.id
            borrow.approved_at = datetime.now()
            borrow.reject_reason = reason
            message = '借阅申请已拒绝'
            notif_title = '借阅申请未获批准'
            notif_content = f'您对档案《{archive_title}》的借阅申请未获批准。'
            if reason:
                notif_content += f'原因：{reason}'
            notif_type = 'danger'

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'借阅审批失败：{str(e)}', 'error')
        return redirect(url_for('archive.borrow_manage'))

    _send_borrow_notification(
        user_id=borrow.borrower_id,
        title=notif_title,
        content=notif_content,
        notif_type=notif_type,
        link=url_for('archive.my_borrows')
    )
    flash(message, 'success' if action == 'approve' else 'warning')
    return redirect(url_for('archive.borrow_manage'))


@archive_bp.route('/borrow/return/<int:borrow_id>', methods=['POST'])
@login_required
def borrow_return(borrow_id):
    """确认归还（集成消息通知）"""
    _, _, _, _, ArchiveBorrow, _, _ = get_models()
    from models import db
    
    borrow = ArchiveBorrow.query.get_or_404(borrow_id)
    
    if borrow.status != 'borrowed':
        return jsonify({'success': False, 'message': '只能归还借出中的档案'})
    
    archive_title = borrow.archive_file.title if borrow.archive_file else '档案'
    borrow.status = 'returned'
    borrow.actual_return_date = datetime.now()

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'档案归还确认失败：{str(e)}', 'error')
        return redirect(url_for('archive.borrow_manage'))
    
    # 推送归还确认通知给借阅人
    _send_borrow_notification(
        user_id=borrow.borrower_id,
        title='档案归还已确认',
        content=f'档案《{archive_title}》的归还已由管理员确认，感谢您的配合。',
        notif_type='info',
        link=url_for('archive.my_borrows')
    )
    
    flash('档案归还确认成功', 'success')
    return redirect(url_for('archive.borrow_manage'))


# ==================== 通知系统 ====================

def _send_borrow_notification(user_id, title, content, notif_type='info', link=None):
    """内部通知推送辅助函数（不阻断主流程）"""
    try:
        from archive_models import ArchiveNotification
        from models import db
        notif = ArchiveNotification(
            user_id=user_id,
            title=title,
            content=content,
            notif_type=notif_type,
            link=link
        )
        db.session.add(notif)
        db.session.commit()
    except Exception:
        db.session.rollback()
        pass  # 通知失败不影响主业务






@archive_bp.route('/notifications')
@login_required
def notifications():
    """我的消息通知列表（分页）"""
    try:
        from archive_models import ArchiveNotification
        page = request.args.get('page', 1, type=int)
        per_page = 20
        pagination = ArchiveNotification.query.filter_by(
            user_id=current_user.id
        ).order_by(ArchiveNotification.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        unread_count = ArchiveNotification.query.filter_by(
            user_id=current_user.id, is_read=False
        ).count()
        return render_template('archive/notifications.html',
                               notifications=pagination.items,
                               pagination=pagination,
                               unread_count=unread_count)
    except Exception:
        return render_template('archive/notifications.html',
                               notifications=[], pagination=None, unread_count=0)


@archive_bp.route('/api/notifications/unread_count')
@login_required
def api_unread_count():
    """获取未读通知数（导航栏轮询接口）"""
    try:
        from archive_models import ArchiveNotification
        count = ArchiveNotification.query.filter_by(
            user_id=current_user.id, is_read=False
        ).count()
        return jsonify({'count': count})
    except Exception:
        return jsonify({'count': 0})


@archive_bp.route('/api/notifications/recent')
@login_required
def api_recent_notifications():
    """获取最近5条通知（导航栏下拉预览）"""
    try:
        from archive_models import ArchiveNotification
        notifs = ArchiveNotification.query.filter_by(
            user_id=current_user.id
        ).order_by(ArchiveNotification.created_at.desc()).limit(5).all()
        return jsonify({
            'notifications': [
                {
                    'id': n.id,
                    'title': n.title,
                    'content': n.content or '',
                    'notif_type': n.notif_type or 'info',
                    'is_read': n.is_read,
                    'link': n.link or '',
                    'created_at': n.created_at.strftime('%m-%d %H:%M') if n.created_at else ''
                }
                for n in notifs
            ]
        })
    except Exception:
        return jsonify({'notifications': []})


@archive_bp.route('/api/notifications/read', methods=['POST'])
@login_required
def api_mark_read():
    """标记通知为已读（单条或全部）"""
    try:
        from archive_models import ArchiveNotification
        from models import db
        import json as _json
        data = request.get_json() or {}
        
        if data.get('all'):
            ArchiveNotification.query.filter_by(
                user_id=current_user.id, is_read=False
            ).update({'is_read': True})
        else:
            notif_id = data.get('id')
            if notif_id:
                notif = ArchiveNotification.query.filter_by(
                    id=notif_id, user_id=current_user.id
                ).first()
                if notif:
                    notif.is_read = True
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


@archive_bp.route('/api/notifications/read_all', methods=['POST'])
@login_required
def api_mark_all_read():
    """标记全部通知为已读"""
    try:
        from archive_models import ArchiveNotification
        from models import db
        ArchiveNotification.query.filter_by(
            user_id=current_user.id, is_read=False
        ).update({'is_read': True})
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})



# ==================== 统计报表 ====================

@archive_bp.route('/statistics')
@login_required
def statistics():
    """档案统计"""
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()
    from models import db
    
    # 总体统计
    stats = {
        'total_files': ArchiveFile.query.count(),
        'digitized': ArchiveFile.query.filter_by(is_digitized=True).count(),
        'permanent': ArchiveFile.query.filter_by(retention_period='永久').count(),
    }
    
    # 保管期限统计
    retention_data = {
        'permanent': ArchiveFile.query.filter_by(retention_period='永久').count(),
        'y30': ArchiveFile.query.filter_by(retention_period='30年').count(),
        'y10': ArchiveFile.query.filter_by(retention_period='10年').count(),
        'unknown': ArchiveFile.query.filter(
            db.or_(
                ArchiveFile.retention_period == None,
                ArchiveFile.retention_period == ''
            )
        ).count(),
    }
    
    # 安全级别统计
    security_data = {
        'public': ArchiveFile.query.filter_by(security_level='公开').count(),
        'internal': ArchiveFile.query.filter_by(security_level='内部').count(),
        'secret': ArchiveFile.query.filter_by(security_level='机密').count(),
        'confidential': ArchiveFile.query.filter_by(security_level='秘密').count(),
    }
    
    # 按类型统计
    type_data = db.session.query(
        ArchiveFile.archive_type,
        db.func.count(ArchiveFile.id).label('count')
    ).group_by(ArchiveFile.archive_type).all()
    type_data = [{'type': t[0] or '未分类', 'count': t[1]} for t in type_data]
    
    # 按年度统计
    year_data = db.session.query(
        ArchiveFile.file_year,
        db.func.count(ArchiveFile.id).label('count')
    ).group_by(ArchiveFile.file_year).order_by(
        ArchiveFile.file_year.desc()
    ).limit(10).all()
    year_data = [{'year': y[0] or '未知', 'count': y[1]} for y in year_data]
    
    return render_template('archive/statistics.html',
                         stats=stats,
                         retention_data=retention_data,
                         security_data=security_data,
                         type_data=type_data,
                         year_data=year_data)


# ==================== 数字化任务 ====================

@archive_bp.route('/digitization/tasks')
@login_required
def digitization_tasks():
    """数字化任务列表"""
    _, _, _, _, _, _, ArchiveDigitizationTask = get_models()
    
    tasks = ArchiveDigitizationTask.query.order_by(
        ArchiveDigitizationTask.created_at.desc()
    ).all()
    return render_template('archive/digitization_tasks.html', tasks=tasks)


@archive_bp.route('/digitization/task/create', methods=['POST'])
@login_required
def digitization_task_create():
    """创建数字化任务"""
    _, _, _, _, _, _, ArchiveDigitizationTask = get_models()
    from models import db
    
    data = request.form
    task = ArchiveDigitizationTask(
        task_name=data.get('task_name'),
        fonds_id=data.get('fonds_id'),
        catalog_id=data.get('catalog_id'),
        year_start=data.get('year_start'),
        year_end=data.get('year_end'),
        retention_period=data.get('retention_period'),
        scan_resolution=data.get('scan_resolution', '300'),
        color_mode=data.get('color_mode', 'color'),
        enable_ocr=data.get('enable_ocr') == 'on',
        created_by=current_user.id
    )
    db.session.add(task)
    db.session.commit()
    flash('数字化任务创建成功', 'success')
    return redirect(url_for('archive.digitization_tasks'))


# ==================== 批量任务管理 ====================

@archive_bp.route('/tasks')
@login_required
def task_list():
    """批量任务列表"""
    from batch_processor import BatchTask
    
    # 获取最近的任务
    recent_tasks = BatchTask.query.order_by(
        BatchTask.created_at.desc()
    ).limit(30).all()
    
    return render_template('archive/task_list.html', 
                         tasks=[t.to_dict() for t in recent_tasks])


@archive_bp.route('/api/task/status/<task_id>')
@login_required
def api_task_status(task_id):
    """获取任务状态"""
    from batch_processor import BatchTask
    
    task = BatchTask.query.filter_by(task_id=task_id).first()
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    
    return jsonify(task.to_dict())


@archive_bp.route('/api/task/cancel/<task_id>', methods=['POST'])
@login_required
def api_task_cancel(task_id):
    """取消任务"""
    from batch_processor import task_queue
    
    success = task_queue.cancel_task(task_id)
    if success:
        return jsonify({'success': True, 'message': '任务已取消'})
    else:
        return jsonify({'success': False, 'message': '无法取消任务'})


@archive_bp.route('/api/task/retry/<task_id>', methods=['POST'])
@login_required
def api_task_retry(task_id):
    """重试任务"""
    from batch_processor import task_queue, create_batch_import_task
    from batch_processor import BatchTask
    
    # 获取原始任务
    old_task = BatchTask.query.filter_by(task_id=task_id).first()
    if not old_task or old_task.status != 'failed':
        return jsonify({'success': False, 'message': '只能重试失败的任务'})
    
    # 创建新任务
    task_data = {
        'file_list': old_task.file_list,
        'fonds_id': old_task.params.get('fonds_id') if old_task.params else None,
        'catalog_id': old_task.params.get('catalog_id') if old_task.params else None,
        'user_id': old_task.created_by
    }
    
    success = task_queue.retry_task(task_id)
    if success:
        return jsonify({'success': True, 'message': '任务已重新提交'})
    else:
        return jsonify({'success': False, 'message': '无法重试任务'})


@archive_bp.route('/api/task/submit', methods=['POST'])
@login_required
def api_task_submit():
    """
    提交批量导入任务
    支持多文件上传
    """
    import os
    import uuid
    from batch_processor import create_batch_import_task
    import utils as utils_module
    from werkzeug.utils import secure_filename
    
    fonds_id = request.form.get('fonds_id')
    catalog_id = request.form.get('catalog_id')
    volume_id = request.form.get('volume_id') or None
    
    if not fonds_id or not catalog_id:
        return jsonify({'error': '请选择全宗和目录'}), 400
    
    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        return jsonify({'error': '请选择要上传的文件'}), 400
    
    # 保存临时文件
    files_data = []
    temp_folder = os.path.join('uploads', 'temp', 'batch_' + datetime.now().strftime('%Y%m%d%H%M%S'))
    os.makedirs(temp_folder, exist_ok=True)
    
    for file in files:
        if file and file.filename and utils_module.allowed_file(file.filename):
            original_name = secure_filename(file.filename)
            temp_path = os.path.join(temp_folder, original_name)
            file.save(temp_path)
            files_data.append({
                'path': temp_path,
                'filename': original_name,
                'original_name': original_name,
                'size': os.path.getsize(temp_path)
            })
    
    if not files_data:
        return jsonify({'error': '没有有效的文件'}), 400
    
    # 创建批量导入任务
    task_id = create_batch_import_task(
        user_id=current_user.id,
        files_data=files_data,
        fonds_id=int(fonds_id),
        catalog_id=int(catalog_id),
        volume_id=int(volume_id) if volume_id else None
    )
    
    return jsonify({
        'success': True,
        'task_id': task_id,
        'message': f'已提交 {len(files_data)} 个文件到后台处理'
    })


@archive_bp.route('/api/task/progress/<task_id>')
@login_required
def api_task_progress(task_id):
    """获取任务进度（轮询接口）"""
    from batch_processor import BatchTask
    
    task = BatchTask.query.filter_by(task_id=task_id).first()
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    
    return jsonify({
        'task_id': task.task_id,
        'status': task.status,
        'progress': task.progress,
        'current_step': task.current_step,
        'total_items': task.total_items,
        'processed_items': task.processed_items,
        'success_items': task.success_items,
        'failed_items': task.failed_items,
        'error_message': task.error_message
    })


# ==================== Excel元数据导入 ====================

@archive_bp.route('/excel_import', methods=['GET', 'POST'])
@login_required
def excel_import():
    """Excel批量元数据导入 - 支持预览确认"""
    from archive_models import ArchiveFile
    from models import db
    
    if request.method == 'POST':
        action = request.form.get('action', 'preview')
        
        # ===== 正式导入 =====
        if action == 'confirm_import':
            import json
            rows_json = request.form.get('rows_data', '[]')
            try:
                rows = json.loads(rows_json)
                success_count = 0
                error_count = 0
                errors = []
                
                for row in rows:
                    try:
                        archive = ArchiveFile(
                            fonds_id=int(row.get('全宗ID', 0)),
                            catalog_id=int(row.get('目录ID', 0)),
                            title=str(row.get('标题', '')),
                            responsibility=str(row.get('责任者', '')),
                            retention_period=str(row.get('保管期限', '30年')),
                            security_level=str(row.get('密级', '公开')),
                            archive_type=str(row.get('档案类型', '其他')),
                            keywords=str(row.get('关键词', '')),
                            reference_number=str(row.get('文号', '')),
                            description=str(row.get('备注', '')),
                            file_code=str(row.get('件号', '')),
                            status='active',
                            created_by=current_user.id
                        )
                        db.session.add(archive)
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(str(e))
                
                db.session.commit()
                
                if error_count > 0:
                    flash(f'成功导入 {success_count} 条记录，失败 {error_count} 条', 'warning')
                else:
                    flash(f'成功导入 {success_count} 条记录', 'success')
                return redirect(url_for('archive.file_list'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'导入失败: {str(e)}', 'error')
                return redirect(request.url)

        
        # ===== 预览解析 =====
        if 'file' not in request.files:
            flash('请选择Excel文件', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('请上传Excel文件', 'error')
            return redirect(request.url)
        
        try:
            import pandas as pd
            import json
            df = pd.read_excel(file)
            
            # 字段映射：支持多种列名
            field_map = {
                '标题': ['标题', 'title', '题名', '文件名'],
                '全宗ID': ['全宗ID', 'fonds_id', '全宗号'],
                '目录ID': ['目录ID', 'catalog_id', '目录号'],
                '责任者': ['责任者', 'responsibility', '形成单位'],
                '保管期限': ['保管期限', 'retention_period', '保存期限'],
                '密级': ['密级', 'security_level', '保密级别'],
                '档案类型': ['档案类型', 'archive_type', '类型'],
                '关键词': ['关键词', 'keywords', '主题词'],
                '文号': ['文号', 'reference_number', '发文字号'],
                '备注': ['备注', 'description', '说明'],
                '件号': ['件号', 'file_code', '档号'],
            }
            
            # 标准化列名
            rename_map = {}
            for std_name, aliases in field_map.items():
                for alias in aliases:
                    if alias in df.columns and alias != std_name:
                        rename_map[alias] = std_name
                        break
            if rename_map:
                df = df.rename(columns=rename_map)
            
            # 验证必要字段
            required_cols = ['标题', '全宗ID', '目录ID']
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                flash(f'缺少必要字段: {", ".join(missing)}，请下载模板参考格式', 'error')
                return redirect(request.url)
            
            # 准备预览数据
            preview_rows = []
            error_rows = []
            
            for idx, row in df.iterrows():
                row_data = {}
                errors = []
                
                for col in field_map.keys():
                    if col in df.columns:
                        val = row.get(col, '')
                        row_data[col] = '' if (val != val) else str(val)  # NaN处理
                    else:
                        row_data[col] = ''
                
                # 验证
                if not row_data.get('标题'):
                    errors.append('标题不能为空')
                if not row_data.get('全宗ID') or row_data['全宗ID'] in ('', 'nan'):
                    errors.append('全宗ID不能为空')
                if not row_data.get('目录ID') or row_data['目录ID'] in ('', 'nan'):
                    errors.append('目录ID不能为空')
                
                row_data['_row_index'] = idx + 2  # Excel行号（从2开始）
                row_data['_errors'] = errors
                
                if errors:
                    error_rows.append(row_data)
                else:
                    preview_rows.append(row_data)
            
            import json as json_module
            return render_template('archive/excel_import.html',
                                 preview_rows=preview_rows,
                                 error_rows=error_rows,
                                 rows_json=json_module.dumps(preview_rows, ensure_ascii=False))
            
        except Exception as e:
            flash(f'解析失败: {str(e)}', 'error')
            return redirect(request.url)
    
    return render_template('archive/excel_import.html',
                          preview_rows=None, error_rows=None, rows_json='[]')


@archive_bp.route('/excel_template')
@login_required
def excel_template_download():
    """下载Excel导入模板"""
    try:
        import openpyxl
        from flask import send_file
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '档案元数据导入模板'
        
        # 表头
        headers = ['标题', '全宗ID', '目录ID', '件号', '责任者', '保管期限', 
                   '密级', '档案类型', '关键词', '文号', '备注']
        header_hints = [
            '档案题名（必填）', '全宗数字ID（必填）', '目录数字ID（必填）',
            '件号（可留空自动生成）', '形成单位/责任者', 
            '永久/30年/10年（默认30年）', '公开/内部/机密/绝密（默认公开）',
            '文书/科技/会计/声像（默认其他）', '关键词逗号分隔', '发文字号', '备注说明'
        ]
        
        # 样式
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        hint_fill = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_idx, (header, hint) in enumerate(zip(headers, header_hints), 1):
            # 第1行：列名
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
            # 第2行：说明
            hint_cell = ws.cell(row=2, column=col_idx, value=hint)
            hint_cell.fill = hint_fill
            hint_cell.font = Font(color='2B6CB0', size=9, italic=True)
            hint_cell.alignment = Alignment(horizontal='center')
        
        # 示例数据行
        sample_data = [
            ['关于加强档案管理工作的通知', '1', '1', '001', '县人民政府办公室', '永久', '公开', '文书', '档案管理,规范化', '县府发〔2024〕1号', '示例数据可删除'],
            ['2024年度工作总结报告', '1', '2', '001', '县档案局', '30年', '内部', '文书', '工作总结,年度报告', '', ''],
        ]
        for row_data in sample_data:
            ws.append(row_data)
        
        # 调整列宽
        col_widths = [30, 10, 10, 12, 20, 12, 12, 12, 25, 20, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # 冻结首行
        ws.freeze_panes = 'A3'
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='档案元数据导入模板.xlsx'
        )
    except Exception as e:
        flash(f'模板生成失败: {str(e)}', 'error')
        return redirect(url_for('archive.excel_import'))


# ==================== 国标质检 API ====================

@archive_bp.route('/api/quality_check/<int:file_id>', methods=['POST'])
@login_required
def api_quality_check(file_id):
    """对单份档案执行质检"""
    archive_digitizer = get_digitizer()
    report = archive_digitizer.run_quality_check_on_archive(file_id)
    return jsonify(report)


@archive_bp.route('/api/quality_check/batch', methods=['POST'])
@login_required
def api_quality_check_batch():
    """批量质检"""
    archive_digitizer = get_digitizer()
    fonds_id = request.form.get('fonds_id', type=int)
    catalog_id = request.form.get('catalog_id', type=int)
    unchecked_only = request.form.get('unchecked_only', 'true').lower() == 'true'

    report = archive_digitizer.run_batch_quality_check(
        fonds_id=fonds_id,
        catalog_id=catalog_id,
        unchecked_only=unchecked_only,
    )
    return jsonify(report)


@archive_bp.route('/api/quality_check/report/<int:file_id>')
@login_required
def api_quality_report(file_id):
    """获取已保存的质检报告"""
    _, _, _, ArchiveFile, _, _, _ = get_models()
    import json
    af = ArchiveFile.query.get_or_404(file_id)
    if af.quality_report:
        try:
            return jsonify(json.loads(af.quality_report))
        except Exception:
            pass
    return jsonify({'error': '暂无质检报告，请先执行质检'}), 404


@archive_bp.route('/quality')
@login_required
def quality_dashboard():
    """质检统计仪表板"""
    ArchiveFonds, _, _, ArchiveFile, _, _, _ = get_models()
    from models import db
    from sqlalchemy import func

    # 整体统计
    total = ArchiveFile.query.filter_by(status='active', is_digitized=True).count()
    checked = ArchiveFile.query.filter_by(status='active', quality_checked=True).count()
    passed = ArchiveFile.query.filter_by(status='active', quality_checked=True).filter(
        ArchiveFile.quality_score >= 80
    ).count()
    dpi_fail = ArchiveFile.query.filter_by(status='active', dpi_compliant=False).count()
    format_fail = ArchiveFile.query.filter_by(status='active', format_compliant=False).count()

    # 分数分布
    score_dist = {
        'A（90-100）': ArchiveFile.query.filter(
            ArchiveFile.status == 'active',
            ArchiveFile.quality_score >= 90
        ).count(),
        'B（75-89）': ArchiveFile.query.filter(
            ArchiveFile.status == 'active',
            ArchiveFile.quality_score.between(75, 89)
        ).count(),
        'C（60-74）': ArchiveFile.query.filter(
            ArchiveFile.status == 'active',
            ArchiveFile.quality_score.between(60, 74)
        ).count(),
        'D（<60）': ArchiveFile.query.filter(
            ArchiveFile.status == 'active',
            ArchiveFile.quality_score < 60
        ).count(),
    }

    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()

    return render_template('archive/quality_dashboard.html',
                         total=total,
                         checked=checked,
                         passed=passed,
                         dpi_fail=dpi_fail,
                         format_fail=format_fail,
                         score_dist=score_dist,
                         fonds_list=fonds_list)


@archive_bp.route('/api/naming/validate', methods=['POST'])
@login_required
def api_validate_naming():
    """验证档号命名格式"""
    from archive_naming import archive_naming
    code = request.form.get('archive_code', '')
    result = archive_naming.validate_archive_code(code)
    return jsonify(result)


@archive_bp.route('/api/naming/generate', methods=['POST'])
@login_required
def api_generate_archive_code():
    """生成标准档号"""
    from archive_naming import archive_naming
    ArchiveFonds, ArchiveCatalog, ArchiveVolume, ArchiveFile, _, _, _ = get_models()

    fonds_id = request.form.get('fonds_id', type=int)
    catalog_id = request.form.get('catalog_id', type=int)
    volume_id = request.form.get('volume_id', type=int)
    retention = request.form.get('retention_period', '30年')
    security = request.form.get('security_level', '公开')

    if not fonds_id or not catalog_id:
        return jsonify({'error': '请选择全宗和目录'}), 400

    fonds = ArchiveFonds.query.get(fonds_id)
    catalog = ArchiveCatalog.query.get(catalog_id)
    volume = ArchiveVolume.query.get(volume_id) if volume_id else None

    # 获取下一个序号
    query = ArchiveFile.query.filter_by(catalog_id=catalog_id)
    if volume_id:
        query = query.filter_by(volume_id=volume_id)
    next_seq = query.count() + 1

    code = archive_naming.generate_archive_code(
        fonds_code=fonds.fonds_code,
        catalog_code=catalog.catalog_code,
        volume_code=volume.volume_code if volume else str(datetime.now().year),
        file_seq=next_seq,
        retention=retention,
        security=security,
    )

    return jsonify({
        'archive_code': code,
        'next_seq': next_seq,
        'validation': archive_naming.validate_archive_code(code),
    })


@archive_bp.route('/api/image/features')
@login_required
def api_image_features():
    """查询图像处理功能可用状态"""
    from archive_image_processor import archive_image_processor
    features = archive_image_processor.get_available_features()
    return jsonify(features)


# ==================== 批量操作 ====================

@archive_bp.route('/batch_delete', methods=['POST'])
@login_required
def batch_delete():
    """批量删除档案"""
    from archive_models import ArchiveFile
    from models import db
    import json
    
    ids = request.form.get('ids')
    if not ids:
        return jsonify({'error': '未选择要删除的档案'}), 400
    
    try:
        file_ids = json.loads(ids)
        count = ArchiveFile.query.filter(ArchiveFile.id.in_(file_ids)).delete(
            synchronize_session=False
        )
        db.session.commit()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@archive_bp.route('/batch_update_status', methods=['POST'])
@login_required
def batch_update_status():
    """批量更新档案状态"""
    from archive_models import ArchiveFile
    from models import db
    import json
    
    ids = request.form.get('ids')
    status = request.form.get('status')
    
    if not ids or not status:
        return jsonify({'error': '参数不完整'}), 400
    
    try:
        file_ids = json.loads(ids)
        count = ArchiveFile.query.filter(ArchiveFile.id.in_(file_ids)).update(
            {'status': status},
            synchronize_session=False
        )
        db.session.commit()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ==================== 数字化管理面板 ====================

@archive_bp.route('/digitize_manage')
@login_required
def digitize_manage():
    """
    数字化管理统一界面
    整合：任务队列 / 最新档案 / 图像处理工具 / 批量质检
    """
    _, _, _, ArchiveFile, _, _, _ = get_models()
    from models import db
    from sqlalchemy import func

    # 统计数据
    total = ArchiveFile.query.filter_by(status='active', is_digitized=True).count()
    checked = ArchiveFile.query.filter_by(status='active', quality_checked=True).count()
    passed = ArchiveFile.query.filter(
        ArchiveFile.status == 'active',
        ArchiveFile.quality_checked == True,
        ArchiveFile.quality_score >= 80
    ).count()
    pending = total - checked
    avg_score = db.session.query(func.avg(ArchiveFile.quality_score)).filter(
        ArchiveFile.status == 'active',
        ArchiveFile.quality_score.isnot(None)
    ).scalar() or 0

    # 最新数字化档案（10条）
    recent_archives = ArchiveFile.query.filter_by(
        status='active', is_digitized=True
    ).order_by(ArchiveFile.digitized_at.desc()).limit(10).all()

    # 全宗列表
    ArchiveFonds, _, _, _, _, _, _ = get_models()
    fonds_list = ArchiveFonds.query.filter_by(is_active=True).all()

    return render_template('archive/digitize_manage.html',
                         stats={
                             'total': total,
                             'passed': passed,
                             'pending': pending,
                             'avg_score': round(avg_score, 1),
                         },
                         recent_archives=recent_archives,
                         fonds_list=fonds_list)


@archive_bp.route('/api/digitization/tasks')
@login_required
def digitization_tasks_api():
    """获取当前用户的后台任务列表（AJAX轮询）"""
    from batch_processor import BatchTask

    tasks = BatchTask.query.filter_by(
        user_id=current_user.id
    ).order_by(BatchTask.created_at.desc()).limit(10).all()

    status_text_map = {
        'pending': '等待中',
        'processing': '处理中',
        'completed': '已完成',
        'failed': '失败',
        'cancelled': '已取消',
    }

    return jsonify({
        'tasks': [{
            'id': t.id,
            'task_id': t.task_id,
            'name': t.task_name or f'任务 #{t.task_id[:8]}',
            'status': t.status,
            'status_text': status_text_map.get(t.status, t.status),
            'progress': t.progress or 0,
            'processed': t.processed_items or 0,
            'total': t.total_items or 0,
        } for t in tasks]
    })


@archive_bp.route('/api/image/process', methods=['POST'])
@login_required
def api_image_process():
    """图像处理工具 API"""
    from archive_image_processor import archive_image_processor

    if 'image_file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['image_file']
    if not file.filename:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    temp_path = save_upload_file(file, 'temp')
    if not temp_path:
        return jsonify({'success': False, 'error': '文件保存失败'}), 500

    try:
        file_code = f"img_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        color_mode = request.form.get('color_mode', 'auto')
        target_dpi = int(request.form.get('target_dpi', 300))
        enable_deskew = request.form.get('enable_deskew') == 'on'
        enable_border = request.form.get('enable_border_removal') == 'on'

        result = archive_image_processor.process_image(
            source_path=temp_path,
            file_code=file_code,
            enable_deskew=enable_deskew,
            enable_border_removal=enable_border,
            target_dpi=target_dpi,
            color_mode=color_mode,
        )

        return jsonify({
            'success': True,
            'quality_score': result.get('quality_score', 0),
            'dpi': result.get('dpi', 0),
            'color_mode': result.get('color_mode', ''),
            'deskew_angle': result.get('deskew_angle', 0),
            'tiff_path': os.path.basename(result.get('tiff_path', '')) if result.get('tiff_path') else None,
            'jpeg_path': os.path.basename(result.get('jpeg_path', '')) if result.get('jpeg_path') else None,
            'pdf_path': os.path.basename(result.get('pdf_path', '')) if result.get('pdf_path') else None,
            'warnings': result.get('warnings', []),
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)




