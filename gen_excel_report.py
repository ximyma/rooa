import sys
sys.stdout.reconfigure(encoding='utf-8')
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date

DB_PATH = r'C:\Users\Administrator\Desktop\ooa\instance\oa.db'
OUTPUT_PATH = r'C:\Users\Administrator\Desktop\ooa\briefing_output\网站栏目监测结果_20260412.xlsx'

# 分类到默认deadline_days的映射
CATEGORY_DEFAULT_DAYS = {
    '年度更新': 365,
    '季度更新': 90,
    '月度更新': 30,
    '动态要闻': 14,
}

conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()
cursor.execute("""
    SELECT column_name, column_category, update_deadline, deadline_days, 
           last_max_date, days_since_update, url, status
    FROM monitor_results 
    WHERE date(monitor_time)=date('now')
    ORDER BY id
""")
rows = cursor.fetchall()
conn.close()

print(f"读取到 {len(rows)} 条记录")

# 处理数据
data = []
for row in rows:
    column_name, column_category, update_deadline, deadline_days, last_max_date, days_since_update, url, status = row
    
    # 补全缺失的deadline_days
    if deadline_days is None or deadline_days == 0:
        cat = column_category or ''
        if '年度' in cat:
            deadline_days = 365
        elif '季度' in cat:
            deadline_days = 90
        elif '月度' in cat:
            deadline_days = 30
        elif '动态' in cat or '要闻' in cat:
            deadline_days = 14
        else:
            deadline_days = 365  # 默认年度
    
    # 实际间隔天数
    actual_days = days_since_update if days_since_update is not None else 0
    
    # 剩余天数
    remaining = deadline_days - actual_days
    
    # 状态
    if remaining <= 0:
        status_text = '已逾期'
    else:
        status_text = f'剩余{remaining}天'
    
    data.append({
        'column_name': column_name or '',
        'last_max_date': str(last_max_date) if last_max_date else '无',
        'deadline_days': deadline_days,
        'actual_days': actual_days,
        'remaining': remaining,
        'status': status_text,
        'column_category': column_category or '',
        'update_deadline': update_deadline or '',
        'url': url or '',
    })

# 按剩余天数升序排列
data.sort(key=lambda x: x['remaining'])

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = '网站栏目监测结果'

# 标题行
headers = ['序号', '栏目名称', '最后更新', '更新期限天数', '实际间隔天数', '剩余天数', '状态', '栏目分类', '更新要求', '网址']
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 颜色定义
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
red_font = Font(color='FFFFFF', bold=True)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

for idx, item in enumerate(data, 1):
    row_num = idx + 1
    values = [
        idx,
        item['column_name'],
        item['last_max_date'],
        item['deadline_days'],
        item['actual_days'],
        item['remaining'],
        item['status'],
        item['column_category'],
        item['update_deadline'],
        item['url'],
    ]
    
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=(col == 2 or col == 10))
        
        # 数字列居中
        if col in (1, 4, 5, 6):
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 按剩余天数设置行颜色
    remaining = item['remaining']
    if remaining <= 0:
        fill = red_fill
        font = red_font
    elif remaining < 8:
        fill = yellow_fill
        font = Font()
    else:
        fill = green_fill
        font = Font()
    
    # 给状态列和剩余天数列上色
    ws.cell(row=row_num, column=6).fill = fill
    ws.cell(row=row_num, column=6).font = font
    ws.cell(row=row_num, column=7).fill = fill
    ws.cell(row=row_num, column=7).font = font

# 设置列宽
col_widths = {
    'A': 6,   # 序号
    'B': 50,  # 栏目名称
    'C': 14,  # 最后更新
    'D': 14,  # 更新期限天数
    'E': 14,  # 实际间隔天数
    'F': 12,  # 剩余天数
    'G': 12,  # 状态
    'H': 14,  # 栏目分类
    'I': 14,  # 更新要求
    'J': 45,  # 网址
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 冻结首行
ws.freeze_panes = 'A2'

# 自动筛选
ws.auto_filter.ref = f'A1:J{len(data)+1}'

wb.save(OUTPUT_PATH)
print(f"Excel报告已保存: {OUTPUT_PATH}")
print(f"总记录数: {len(data)}")
