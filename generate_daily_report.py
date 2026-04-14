# -*- coding: utf-8 -*-
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

conn = sqlite3.connect('instance/oa.db')
cursor = conn.cursor()

# 获取今天的记录
cursor.execute('''
    SELECT id, column_name, last_max_date, deadline_days, days_since_update,
           column_category, update_deadline, url, status, is_overdue, is_expiring
    FROM monitor_results 
    WHERE date(monitor_time) = date('now')
    ORDER BY days_since_update DESC
''')
rows = cursor.fetchall()
print(f'Found {len(rows)} records for today')

# deadline_days 映射
def get_deadline_days(row):
    deadline_days = row[3]
    if deadline_days is not None:
        return deadline_days
    category = row[5]
    deadline_str = row[6]
    if category and ('年度' in str(category) or '年度' in str(deadline_str)):
        return 365
    elif category and ('季度' in str(category) or '季度' in str(deadline_str)):
        return 90
    elif category and ('月度' in str(category) or '月度' in str(deadline_str)):
        return 30
    elif category and ('动态' in str(category) or '动态' in str(deadline_str) or '要闻' in str(category)):
        return 14
    return 365

# 准备数据
data = []
today = date.today()
for i, row in enumerate(rows, 1):
    deadline_days = get_deadline_days(row)
    actual_days = row[4] if row[4] else 0
    remaining = deadline_days - actual_days
    
    if remaining <= 0:
        status = '已逾期'
    elif remaining < 8:
        status = f'剩余{remaining}天'
    else:
        status = f'剩余{remaining}天'
    
    last_update = row[2] if row[2] else ''
    column_name = row[1] if row[1] else ''
    category = row[5] if row[5] else ''
    url = row[7] if row[7] else ''
    deadline_str = row[6] if row[6] else ''
    
    data.append({
        '序号': i,
        '栏目名称': column_name,
        '最后更新': str(last_update) if last_update else '',
        '更新期限天数': deadline_days,
        '实际间隔天数': actual_days,
        '剩余天数': remaining,
        '状态': status,
        '栏目分类': category,
        '更新要求': deadline_str,
        '网址': url
    })

# 创建Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '网站栏目监测结果'

# 样式定义
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
yellow_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
green_fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['序号', '栏目名称', '最后更新', '更新期限天数', '实际间隔天数', '剩余天数', '状态', '栏目分类', '更新要求', '网址']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 数据行 - 按剩余天数升序排列
data.sort(key=lambda x: x['剩余天数'])
for row_idx, item in enumerate(data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=item[header])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 设置状态颜色
    remaining = item['剩余天数']
    if remaining <= 0:
        fill = red_fill
    elif remaining < 8:
        fill = yellow_fill
    else:
        fill = green_fill
    
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

# 设置列宽
widths = [8, 40, 12, 12, 12, 10, 12, 15, 12, 50]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 保存
output_dir = 'briefing_output'
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, '网站栏目监测结果_' + today.strftime('%Y%m%d') + '.xlsx')
wb.save(output_file)
print(f'Excel report generated: {output_file}')
print(f'Total records: {len(data)}')

# 统计
overdue = sum(1 for d in data if d['剩余天数'] <= 0)
expiring = sum(1 for d in data if 0 < d['剩余天数'] < 8)
ok = sum(1 for d in data if d['剩余天数'] >= 8)
print(f'Overdue: {overdue}, Expiring: {expiring}, OK: {ok}')

conn.close()
