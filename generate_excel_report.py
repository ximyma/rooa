# -*- coding: utf-8 -*-
"""生成网站栏目监测Excel报告"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import sqlite3
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 数据库路径
DB_PATH = r'C:\Users\Administrator\Desktop\ooa\instance\oa.db'
OUTPUT_DIR = r'C:\Users\Administrator\Desktop\ooa\briefing_output'

# 分类对应期限天数映射
CATEGORY_DEADLINE_MAP = {
    '年度更新': 365,
    '季度更新': 90,
    '月度更新': 30,
    '动态要闻': 14,
}

def get_deadline_days(category, original_deadline):
    """获取期限天数，如果原值为空则根据分类自动补全"""
    if original_deadline is not None and original_deadline > 0:
        return original_deadline
    
    # 根据分类自动补全
    for cat_key, days in CATEGORY_DEADLINE_MAP.items():
        if cat_key in (category or ''):
            return days
    
    # 默认30天
    return 30

def get_status_text(remaining_days):
    """获取状态文本"""
    if remaining_days <= 0:
        return '已逾期'
    else:
        return f'剩余{remaining_days}天'

def main():
    # 连接数据库
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 查询今日监测记录
    today = date.today().strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT 
            column_name,
            last_max_date,
            deadline_days,
            days_since_update,
            column_category,
            update_deadline,
            url
        FROM monitor_results
        WHERE date(monitor_time) = date('now')
        ORDER BY id
    ''')
    
    rows = cursor.fetchall()
    print(f"查询到 {len(rows)} 条今日监测记录")
    
    # 准备数据
    data = []
    for row in rows:
        deadline = get_deadline_days(row['column_category'], row['deadline_days'])
        days_since = row['days_since_update'] or 0
        remaining = deadline - days_since
        
        data.append({
            'column_name': row['column_name'] or '',
            'last_update': row['last_max_date'] or '无',
            'deadline_days': deadline,
            'actual_days': days_since,
            'remaining_days': remaining,
            'status': get_status_text(remaining),
            'category': row['column_category'] or '',
            'requirement': row['update_deadline'] or '',
            'url': row['url'] or ''
        })
    
    # 按剩余天数升序排列
    data.sort(key=lambda x: x['remaining_days'])
    
    conn.close()
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '网站栏目监测结果'
    
    # 定义样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 状态颜色
    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
    green_fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['序号', '栏目名称', '最后更新', '更新期限天数', '实际间隔天数', '剩余天数', '状态', '栏目分类', '更新要求', '网址']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    col_widths = [6, 50, 12, 12, 12, 10, 10, 20, 20, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 写入数据
    for idx, row_data in enumerate(data, 1):
        row_num = idx + 1
        
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=2, value=row_data['column_name']).border = thin_border
        ws.cell(row=row_num, column=3, value=row_data['last_update']).border = thin_border
        ws.cell(row=row_num, column=4, value=row_data['deadline_days']).border = thin_border
        ws.cell(row=row_num, column=5, value=row_data['actual_days']).border = thin_border
        
        remaining_cell = ws.cell(row=row_num, column=6, value=row_data['remaining_days'])
        remaining_cell.border = thin_border
        
        status_cell = ws.cell(row=row_num, column=7, value=row_data['status'])
        status_cell.border = thin_border
        
        # 设置状态颜色
        remaining = row_data['remaining_days']
        if remaining <= 0:
            remaining_cell.fill = red_fill
            status_cell.fill = red_fill
        elif remaining < 8:
            remaining_cell.fill = yellow_fill
            status_cell.fill = yellow_fill
        else:
            remaining_cell.fill = green_fill
            status_cell.fill = green_fill
        
        ws.cell(row=row_num, column=8, value=row_data['category']).border = thin_border
        ws.cell(row=row_num, column=9, value=row_data['requirement']).border = thin_border
        ws.cell(row=row_num, column=10, value=row_data['url']).border = thin_border
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for row_num in range(2, len(data) + 2):
        ws.row_dimensions[row_num].height = 20
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    today_str = datetime.now().strftime('%Y%m%d')
    output_path = f'{OUTPUT_DIR}\\网站栏目监测结果_{today_str}.xlsx'
    wb.save(output_path)
    print(f"Excel报告已生成: {output_path}")
    print(f"共 {len(data)} 条记录")

if __name__ == '__main__':
    main()
