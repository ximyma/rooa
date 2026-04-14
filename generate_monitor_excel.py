# -*- coding: utf-8 -*-
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os

DB_PATH = 'instance/oa.db'
OUTPUT_DIR = 'briefing_output'
today_str = date.today().strftime('%Y%m%d')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'网站栏目监测结果_{today_str}.xlsx')

# Ensure output dir exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

conn = sqlite3.connect(DB_PATH)
c = conn.cursor()

# Get all today's records
c.execute("""
    SELECT column_name, last_max_date, deadline_days, days_since_update,
           column_category, update_deadline, url
    FROM monitor_results
    WHERE date(monitor_time) = date('now')
""")

rows = c.fetchall()
conn.close()

print(f"Total records: {len(rows)}")

# Process data
data = []
for row in rows:
    col_name, last_date, deadline_days, days_since, category, update_req, url = row
    
    # Auto-fill deadline_days based on category
    if deadline_days is None:
        cat = category or ''
        if '年度' in cat:
            deadline_days = 365
        elif '季度' in cat:
            deadline_days = 90
        elif '月度' in cat:
            deadline_days = 30
        elif '动态要闻' in cat or '动态' in cat:
            deadline_days = 14
        else:
            deadline_days = 365  # default
    
    # Calculate remaining days
    remaining = deadline_days - days_since
    
    # Status
    if remaining <= 0:
        status = f"已逾期{abs(remaining)}天"
    else:
        status = f"剩余{remaining}天"
    
    data.append({
        'column_name': col_name or '',
        'last_date': last_date or '',
        'deadline_days': deadline_days,
        'days_since': days_since or 0,
        'remaining': remaining,
        'status': status,
        'category': category or '',
        'update_req': update_req or '',
        'url': url or ''
    })

# Sort by remaining days ascending
data.sort(key=lambda x: x['remaining'])

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '网站栏目监测结果'

# Styles
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='微软雅黑', size=10)
cell_alignment = Alignment(vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
yellow_fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')
green_fill = PatternFill(start_color='69DB7C', end_color='69DB7C', fill_type='solid')
red_font = Font(name='微软雅黑', size=10, color='FFFFFF', bold=True)
yellow_font = Font(name='微软雅黑', size=10, color='333333')
green_font = Font(name='微软雅黑', size=10, color='006600')

# Headers
headers = ['序号', '栏目名称', '最后更新', '更新期限天数', '实际间隔天数', '剩余天数', '状态', '栏目分类', '更新要求', '网址']
col_widths = [8, 50, 14, 14, 14, 12, 18, 20, 20, 50]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

# Data rows
for idx, item in enumerate(data, 1):
    row_num = idx + 1
    values = [
        idx,
        item['column_name'],
        str(item['last_date']),
        item['deadline_days'],
        item['days_since'],
        item['remaining'],
        item['status'],
        item['category'],
        item['update_req'],
        item['url']
    ]
    
    remaining = item['remaining']
    
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = thin_border
        cell.font = cell_font
        
        # Color coding based on remaining days (status column and remaining column)
        if remaining <= 0:
            if col_idx in (6, 7):  # remaining and status columns
                cell.fill = red_fill
                cell.font = red_font
            cell.alignment = center_alignment if col_idx in (1, 3, 4, 5, 6, 7) else cell_alignment
        elif remaining < 8:
            if col_idx in (6, 7):
                cell.fill = yellow_fill
                cell.font = yellow_font
            cell.alignment = center_alignment if col_idx in (1, 3, 4, 5, 6, 7) else cell_alignment
        else:
            if col_idx in (6, 7):
                cell.fill = green_fill
                cell.font = green_font
            cell.alignment = center_alignment if col_idx in (1, 3, 4, 5, 6, 7) else cell_alignment
        
        if col_idx not in (6, 7):
            if col_idx in (1, 3, 4, 5):
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment

# Freeze top row
ws.freeze_panes = 'A2'

# Auto filter
ws.auto_filter.ref = f"A1:J{len(data)+1}"

wb.save(OUTPUT_FILE)
print(f"Excel saved: {OUTPUT_FILE}")
print(f"Total rows: {len(data)}")

# Count stats
overdue = sum(1 for d in data if d['remaining'] <= 0)
warning = sum(1 for d in data if 0 < d['remaining'] < 8)
ok = sum(1 for d in data if d['remaining'] >= 8)
print(f"Overdue (red): {overdue}")
print(f"Warning (yellow): {warning}")
print(f"OK (green): {ok}")
